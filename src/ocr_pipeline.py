import streamlit as st
import pandas as pd
from PIL import Image
import json
import os
import tempfile
import re
import io
import time
import openpyxl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
from google import genai
from google.genai import types

# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="AMP-GEN Material Passport", layout="wide")

# --- SESSION STATE INITIALIZATION ---
# This is what keeps results alive across reruns (download clicks, slider moves, tab switches).
if "extraction_done" not in st.session_state:
    st.session_state.extraction_done = False
    st.session_state.excel_bytes = None
    st.session_state.json_data = None
    st.session_state.img_bytes = None
    st.session_state.df_extracted = None

# Fallback Models list matching main.py backend logic
FALLBACK_MODELS = [
    "gemini-3.6-flash",
    "gemini-3.7-flash",
    "gemini-3.5-flash",
    "gemini-flash-latest",
    "gemini-3.1-flash-lite"
]

meta_prompt = """
Analyze the first page of this uploaded document.
Extract the building metadata and return EXACTLY a JSON object with these keys:
- Depth_of_Foundation
- Plinth_Height
- Plinth_Area
- No_of_Items
- Seismic_Zone
- Bearing_Capacity
"""

extract_prompt_template = """
You are an expert Quantity Surveyor and Data Engineer. Read the Bill of Quantities PDF. 
Extract ONLY the line items numbered from {start} to {end} inclusive into a structured JSON array of objects. 

CRITICAL RULE FOR SUB-ITEMS: If an item has a main parent description but lists specific sub-items (e.g., i, ii, a, b) with their own quantities below it, you MUST split them into separate JSON objects. Combine the parent description and the child description. Set the "BOQ Item No." to reflect the split (e.g., "31.i", "31.ii"). Do NOT lump multiple quantities into a single row.

Map the extracted data to these specific keys exactly:
- "BOQ Item No.": The serial number (e.g., "1.", "31.i").
- "Description": The full text description of the work (parent + child if split).
- "Original Quantity": The quantity number (e.g., 32.0). If empty, use null.
- "Original Unit": The unit of measurement (e.g., "Cu.m", "Sq.m", "Quintal").
- "Schedule Item Code": The DSR 1989 Code No. (e.g., "2.8"). If empty, use null.
- "Material Category": Infer the primary material (e.g., "Concrete", "Earthwork", "Steel", "Brick", "Wood").
- "Discipline": Infer one of: "Civil & Sitework", "Structural", "Architectural".
- "Floor / Section": Identify the heading/sub-head from the BOQ.
- "Material / Product": A short specific name (e.g., "Ready mix concrete", "TMT Bar").
- "All Materials Detected": Array of strings of all materials mentioned.
- "Material Confidence": "High", "Medium", or "Low".
- "Grade": Any material grade mentioned.
- "Mix Ratio": Any mix ratio mentioned.
- "Standard / Code Reference": Any Indian Standard codes explicitly mentioned.
- "Classification (Matched)": Create a hierarchical classification string.
- "Length (mm)": Extract from description if present (numeric only).
- "Width (mm)": Extract from description if present (numeric only).
- "Height (mm)": Extract from description if present (numeric only).
- "Thickness (mm)": Extract from description if present (e.g., "12 mm thick" -> 12).
- "Depth (mm)": Extract from description if present (numeric only).
- "Diameter (mm)": Extract from description if present (numeric only).
- "Unit Rate": The rate amount if visible, else null.
- "Total Cost": The total amount if visible, else null.
- "Audit Tag": Use "[OK]" for clean extractions, "[FLAG]" for messy/unclear scans, "[EXCLUDED]" for earth/labour, "[MULTI-SUB]" if you successfully split a parent item.

Respond ONLY with the JSON array. Do NOT extract items outside the range {start} to {end}.
"""

# --- HELPER FUNCTIONS ---
def generate_with_fallback(client, contents):
    for model_name in FALLBACK_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=contents,
                config=types.GenerateContentConfig(response_mime_type="application/json")
            )
            return response, model_name
        except Exception as e:
            err_str = str(e)
            if "503" in err_str or "404" in err_str or "429" in err_str:
                continue
            else:
                raise e
    raise RuntimeError("All fallback models are currently unavailable or overloaded.")

def safe_float(val):
    if val is None or val == "": return None
    try:
        clean = re.sub(r'[^\d.]', '', str(val))
        return float(clean) if clean else None
    except: return None

def normalize_unit(raw_unit):
    if not raw_unit: return ""
    u = str(raw_unit).lower().strip()
    if any(x in u for x in ["cum", "cu.m", "m3", "cubic metre", "cubic meter"]): return "cum"
    if any(x in u for x in ["sqm", "sq.m", "m2", "square"]): return "sqm"
    if any(x in u for x in ["rm", "mtr", "metre", "meter"]) or u == "m": return "m"
    if any(x in u for x in ["kg", "kilogram"]): return "kg"
    if any(x in u for x in ["each", "nos", "number"]): return "nos"
    if "quintal" in u: return "quintal"
    if "tonne" in u or "mt" in u: return "tonne"
    return u

# --- UI APP LAYOUT ---
st.title("🏗️ AMP-GEN: Automated Material Passport Generation")
st.markdown("Developed for the AMP-GEN take-home task. This dashboard displays the extracted BoQ data and allows for a live full-extraction demo.")

tab1, tab2, tab3 = st.tabs(["📊 Final Material Passport", "📈 Visualization", "⚡ Live API Demo"])

with tab1:
    st.header("Extracted Material Passport (Excel Data)")
    try:
        df = pd.read_excel("output/passport_filled.xlsx", sheet_name="Material Passport", skiprows=2)
        st.dataframe(df, use_container_width=True)
    except Exception as e:
        st.error(f"Could not load Excel file. Make sure the output folder exists and is populated. Error: {e}")

with tab2:
    st.header("Material Distribution")
    try:
        image = Image.open('output/visualization.png')
        st.image(image, caption='Distribution of Materials across BoQ Items')
    except Exception as e:
        st.error("Visualization not found. Run the live pipeline first to generate it.")

with tab3:
    st.header("⚡ Live Full Pipeline Execution")
    st.markdown("Upload the CBRI BOQ PDF and an optional Excel template to extract, clean, map, and generate deliverables live.")
    
    # --- BATCHES TIME-SAVER SLIDER ---
    st.info("💡 **Evaluator Note:** Full extraction takes ~4-5 minutes due to API rate limits. Use the slider below to process only a selected number of batches.")
    max_batches = st.slider("Number of batches to process (1 batch = ~15 seconds)", min_value=1, max_value=10, value=10)
    
    # Side-by-side file uploaders
    col1, col2 = st.columns(2)
    with col1:
        uploaded_pdf = st.file_uploader("Upload a BoQ PDF document (Required)", type=["pdf"])
    with col2:
        uploaded_template = st.file_uploader("Upload a Custom Excel Template (Optional)", type=["xlsx"])
        st.caption("ℹ️ **Note:** If you leave this blank, the system will automatically use the default AMP Material Passport template.")
    
    api_key_input = st.text_input("Enter Gemini API Key (or leave blank to use server secrets):", type="password")
    
    if st.button("Start Extraction Process", type="primary"):
        raw_keys = api_key_input if api_key_input else st.secrets.get("GEMINI_API_KEYS", "")
        api_keys = [k.strip() for k in raw_keys.split(",") if k.strip()]
        
        if not api_keys:
            st.error("Please provide a valid Gemini API key.")
        elif not uploaded_pdf:
            st.error("Please attach a BoQ PDF document first!")
        else:
            with st.spinner("Preparing temporary files for processing..."):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                    tmp.write(uploaded_pdf.getvalue())
                    tmp_path = tmp.name

            try:
                all_extracted_items = []
                meta_data = {}
                metadata_done = False
                
                # All master chunks mapping to main.py
                all_chunks = [(1, 7), (8, 14), (15, 21), (22, 28), (29, 35), (36, 42), (43, 49), (50, 56), (57, 61), (62, 64)]
                chunks_to_process = all_chunks[:max_batches]
                
                progress_text = f"Batches Done: 0/{len(chunks_to_process)} | Items Extracted: 0"
                progress_bar = st.progress(0, text=progress_text)
                status_text = st.empty()
                
                success_execution = False
                
                for kidx, key in enumerate(api_keys):
                    if metadata_done and len(all_extracted_items) >= (len(chunks_to_process) * 3): 
                        success_execution = True
                        break
                        
                    try:
                        client = genai.Client(api_key=key)
                        boq_file = client.files.upload(file=tmp_path, config={'display_name': f'BoQ_Streamlit_{kidx}'})
                        
                        # 1. Metadata Extraction
                        if not metadata_done:
                            status_text.info("🔍 Extracting Building Metadata from document header...")
                            meta_response, _ = generate_with_fallback(client, [boq_file, meta_prompt])
                            meta_data = json.loads(meta_response.text)
                            metadata_done = True
                            time.sleep(3)
                        
                        # 2. Chunked Line Item Extraction (Driven by Slider)
                        for i, (start, end) in enumerate(chunks_to_process):
                            status_text.info(f"⏳ **Extracting batch {i+1} of {len(chunks_to_process)}:** Items {start} to {end}...")
                            
                            chunk_prompt = extract_prompt_template.format(start=start, end=end)
                            boq_response, _ = generate_with_fallback(client, [boq_file, chunk_prompt])
                            chunk_items = json.loads(boq_response.text)
                            
                            all_extracted_items.extend(chunk_items)
                            
                            progress = (i + 1) / len(chunks_to_process)
                            updated_text = f"Batches Done: {i+1}/{len(chunks_to_process)} | Items Extracted: {len(all_extracted_items)}"
                            progress_bar.progress(progress, text=updated_text)
                            
                            if i < len(chunks_to_process) - 1:
                                time.sleep(10) # rate limit cooldown
                            
                        success_execution = True
                        break
                    except Exception as loop_err:
                        status_text.warning(f"⚠️ Key {kidx+1} hit a limit or error: {loop_err}. Rotating...")
                        continue

                if not success_execution:
                    raise RuntimeError("All configured API keys failed or were exhausted during execution.")

                status_text.success(f"🎉 Extraction Complete ({len(chunks_to_process)} batches)! Cleaning data and mapping to Excel...")

                # --- DATA CLEANING & REGEX ---
                current_floor = ""
                grade_map = {
                    "1:5:10": "M5", "1:4:8": "M7.5", "1:3:6": "M10", "1:2:4": "M15", 
                    "1:1.5:3": "M20", "1:1:2": "M25", "1:6": "Mortar 1:6",
                    "1:5": "Mortar 1:5", "1:4": "Mortar 1:4", "1:3": "Mortar 1:3", "1:2": "Mortar 1:2"
                }

                for item in all_extracted_items:
                    raw_u = str(item.get("Original Unit") or "").lower().strip()
                    qty = safe_float(item.get("Original Quantity"))
                    description = str(item.get("Description") or "")
                    
                    floor = str(item.get("Floor / Section") or "").strip()
                    bad_headers = ["bill of quantities", "p/gen", "modified", "null", "none", ""]
                    if floor and not any(bad in floor.lower() for bad in bad_headers):
                        current_floor = floor 
                    item["Floor / Section"] = current_floor

                    if "decimetre" in raw_u:
                        qty = qty / 1000.0 if qty else 0
                        unit = "cum"
                    else:
                        unit = normalize_unit(raw_u)
                        
                    item["Original Quantity"] = qty
                    item["Original Unit"] = unit

                    is_matches = re.findall(r'(IS\s*[:\-]?\s*\d+)', description, re.IGNORECASE)
                    if is_matches:
                        cleaned_codes = [re.sub(r'IS\s*[:\-]?\s*', 'IS: ', m.upper()) for m in is_matches]
                        item["Standard / Code Reference"] = ", ".join(set(cleaned_codes))

                    mix_ratio = str(item.get("Mix Ratio") or "")
                    grade_match = re.search(r'\b(M\s*[-]?\s*\d{1,2}(?:\.5)?)\b', description, re.IGNORECASE)
                    
                    if grade_match:
                        item["Grade"] = grade_match.group(1).upper().replace(' ', '').replace('-', '')
                    elif mix_ratio and mix_ratio.lower() not in ["null", "none", ""]:
                        ratios = [r.strip().replace(" ", "") for r in mix_ratio.split(',')]
                        mapped_grades = [grade_map[r] for r in ratios if r in grade_map]
                        if mapped_grades:
                            item["Grade"] = ", ".join(mapped_grades)

                # --- EXCEL MAPPING ---
                if uploaded_template:
                    wb = openpyxl.load_workbook(uploaded_template)
                    sheet_name = "Material Passport" if "Material Passport" in wb.sheetnames else wb.sheetnames[0]
                    ws = wb[sheet_name]
                else:
                    wb = openpyxl.load_workbook("data/AMP_Passport_Template.xlsx")
                    ws = wb["Material Passport"]

                if ws.max_row >= 4: ws.delete_rows(4, ws.max_row - 3)
                headers = [cell.value for cell in ws[3]]

                for item in all_extracted_items:
                    qty = item.get("Original Quantity")
                    unit = item.get("Original Unit")
                    mat_cat = str(item.get("Material Category", "")).lower()
                    item_no = str(item.get("BOQ Item No.", "")).strip()
                    
                    gmap_id = f"AMP-CBRI-PR-{item_no.replace('.', '-')}" if item_no else ""
                    if gmap_id.endswith('-'): gmap_id = gmap_id[:-1]

                    all_mats_raw = item.get("All Materials Detected", [])
                    all_mats = ", ".join(str(m) for m in all_mats_raw) if isinstance(all_mats_raw, list) else str(all_mats_raw)
                    
                    density, gwp, db_citation = "", "", ""
                    if "concrete" in mat_cat: density, gwp, db_citation = 2400, 0.15, "ICE Database V3.0"
                    elif "steel" in mat_cat or "reinforcement" in mat_cat: density, gwp, db_citation = 7850, 2.50, "ICE Database V3.0"
                    elif "brick" in mat_cat: density, gwp, db_citation = 1900, 0.24, "ICE Database V3.0"
                    elif "wood" in mat_cat or "timber" in mat_cat: density, gwp, db_citation = 650, 0.45, "ICE Database V3.0"
                    elif "earth" in mat_cat or "sand" in mat_cat: density, gwp, db_citation = 1600, 0.01, "ICE Database V3.0"

                    audit_tag = str(item.get("Audit Tag") or "[OK]")
                    final_comment = f"{audit_tag} negligible carbon" if "EXCLUDED" in audit_tag.upper() else f"{audit_tag} {db_citation}".strip()
                    thick_mm = safe_float(item.get("Thickness (mm)"))

                    row_dict = {
                        "GMAP Id": gmap_id,
                        "BOQ Item No.": item_no,
                        "Description": item.get("Description"),
                        "Original Quantity": qty,
                        "Original Unit": unit,
                        "Schedule Item Code": item.get("Schedule Item Code"),
                        "Material Category": item.get("Material Category"),
                        "Discipline": item.get("Discipline"),
                        "Floor / Section": item.get("Floor / Section"),
                        "Material / Product": item.get("Material / Product"),
                        "All Materials Detected": all_mats,
                        "Material Confidence": item.get("Material Confidence"),
                        "Grade": item.get("Grade"),
                        "Mix Ratio": item.get("Mix Ratio"),
                        "Standard / Code Reference": item.get("Standard / Code Reference"),
                        "Classification (Matched)": item.get("Classification (Matched)"),
                        "Length (mm)": safe_float(item.get("Length (mm)")),
                        "Width (mm)": safe_float(item.get("Width (mm)")),
                        "Height (mm)": safe_float(item.get("Height (mm)")),
                        "Thickness (mm)": thick_mm,
                        "Depth (mm)": safe_float(item.get("Depth (mm)")),
                        "Diameter (mm)": safe_float(item.get("Diameter (mm)")),
                        "Unit Rate": item.get("Unit Rate"),
                        "Total Cost": item.get("Total Cost"),
                        "Currency": "Rs.",
                        "Schedule (DSR/SOR)": "DSR 1989", 
                        "Density (kg/m³)": density,
                        "GWP / kg (kg CO₂e/kg)": gwp,
                        "Comment": final_comment
                    }

                    vol_m3, area_m2, weight_kg = "", "", ""
                    derived_q, derived_u, derived_b = "", "", ""

                    if qty and float(qty) > 0:
                        q = float(qty)
                        if unit == "cum":
                            vol_m3 = q
                            row_dict["Volume (m³)"] = q
                        elif unit == "sqm":
                            area_m2 = q
                            row_dict["Area (m²)"] = q
                            if thick_mm:
                                derived_q = round(q * (thick_mm / 1000.0), 3)
                                derived_u = "Volume (m³)"
                                derived_b = f"Area ({q}) * Thickness ({thick_mm}mm)"
                                vol_m3 = derived_q 
                        elif unit == "m": row_dict["Length (m)"] = q
                        elif unit == "kg": weight_kg = q; row_dict["Weight (kg)"] = q
                        elif unit == "quintal": weight_kg = q * 100; row_dict["Weight (kg)"] = weight_kg
                        elif unit == "tonne": weight_kg = q * 1000; row_dict["Weight (kg)"] = weight_kg
                        elif unit == "nos": row_dict["Count (Nos)"] = q

                    row_dict["Derived Quantity"] = derived_q
                    row_dict["Derived Quantity Unit"] = derived_u
                    row_dict["Derived Quantity Basis"] = derived_b

                    if vol_m3 and density and gwp: row_dict["Embodied Carbon A1-A3 (kg CO₂e)"] = round(vol_m3 * density * gwp, 2)
                    elif weight_kg and gwp: row_dict["Embodied Carbon A1-A3 (kg CO₂e)"] = round(weight_kg * gwp, 2)

                    row_values = [row_dict.get(col, "") for col in headers]
                    ws.append(row_values)
                    
                    for cell in ws[ws.max_row]:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

                # --- PREPARE OUTPUT BYTES ---
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                excel_bytes = excel_buffer.getvalue()
                
                json_data = json.dumps(all_extracted_items, indent=4)
                
                df_extracted = pd.DataFrame(all_extracted_items)
                fig, ax = plt.subplots(figsize=(8, 6))
                if 'Material Category' in df_extracted.columns and not df_extracted['Material Category'].isna().all():
                    cat_counts = df_extracted['Material Category'].value_counts()
                    cat_counts.plot(kind='bar', ax=ax, color='#2ecc71')
                    ax.set_title('Building-Level Material Distribution', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Material Category', fontsize=12)
                    ax.set_ylabel('Number of BOQ Items', fontsize=12)
                    plt.xticks(rotation=45, ha='right')
                    plt.tight_layout()
                else:
                    ax.text(0.5, 0.5, 'No Material Categories Extracted', ha='center', va='center', fontsize=14)
                
                img_buffer = io.BytesIO()
                plt.savefig(img_buffer, format='png', dpi=300)
                img_buffer.seek(0)
                img_bytes = img_buffer.getvalue()
                plt.close(fig)

                # --- SAVE TO output/ SO TAB 1 & TAB 2 PICK THEM UP TOO ---
                os.makedirs("output", exist_ok=True)
                with open("output/passport_filled.xlsx", "wb") as f:
                    f.write(excel_bytes)
                with open("output/passport.json", "w", encoding="utf-8") as f:
                    f.write(json_data)
                with open("output/visualization.png", "wb") as f:
                    f.write(img_bytes)

                # --- PERSIST RESULTS ACROSS RERUNS ---
                st.session_state.excel_bytes = excel_bytes
                st.session_state.json_data = json_data
                st.session_state.img_bytes = img_bytes
                st.session_state.df_extracted = df_extracted
                st.session_state.extraction_done = True

            except Exception as e:
                st.error(f"Error during execution: {e}")
            finally:
                if 'tmp_path' in locals() and os.path.exists(tmp_path):
                    os.remove(tmp_path)

    # --- RESULTS DISPLAY (persists across reruns: downloads, tab switches, slider moves) ---
    if st.session_state.extraction_done:
        st.success("✅ Full extraction, cleaning, and mapping pipeline complete!")

        dl_col1, dl_col2, dl_col3 = st.columns(3)
        with dl_col1:
            st.download_button(
                "📥 Download Excel (.xlsx)",
                data=st.session_state.excel_bytes,
                file_name="passport_filled.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_excel"
            )
        with dl_col2:
            st.download_button(
                "📥 Download JSON (.json)",
                data=st.session_state.json_data,
                file_name="passport.json",
                mime="application/json",
                use_container_width=True,
                key="dl_json"
            )
        with dl_col3:
            st.download_button(
                "📥 Download Chart (.png)",
                data=st.session_state.img_bytes,
                file_name="visualization.png",
                mime="image/png",
                use_container_width=True,
                key="dl_png"
            )

        st.dataframe(st.session_state.df_extracted, use_container_width=True)