import os
import json
import time
import re
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Alignment
from google import genai
from google.genai import types

# ---------------------------------------------------------
# 1. SETUP & CONFIGURATION
# ---------------------------------------------------------
keys_env = os.environ.get("GEMINI_API_KEYS")
if not keys_env:
    raise ValueError("Please set the GEMINI_API_KEYS environment variable (comma-separated).")

api_keys = [k.strip() for k in keys_env.split(",") if k.strip()]

PDF_PATH = "data/BoQ_CBRI_Principals_Residence.pdf"
TEMPLATE_PATH = "data/AMP_Passport_Template.xlsx"

# Ranked list of Flash models for fallback rotation
FALLBACK_MODELS = [
    "gemini-3.6-flash",
    "gemini-3.7-flash",
    "gemini-3.5-flash",
    "gemini-flash-latest",
    "gemini-3.1-flash-lite"
]

# FIX 3: Strict Metadata Schema matching the Instructions sheet
meta_prompt = """
Analyze the first page of this uploaded document.
Extract the building metadata and return EXACTLY a JSON object with these keys:
- Depth_of_Foundation
- Plinth_Height
- Plinth_Area
- No_of_Items
- Seismic_Zone
- Bearing_Capacity
"""

# ENHANCED PROMPT: Handles Sub-items, Dimensions, and Audit Tags
extract_prompt_template = """
You are an expert Quantity Surveyor and Data Engineer. Read the Bill of Quantities PDF. 
Extract ONLY the line items numbered from {start} to {end} inclusive into a structured JSON array of objects. 

CRITICAL RULE FOR SUB-ITEMS: If an item has a main parent description but lists specific sub-items (e.g., i, ii, a, b) with their own quantities below it, you MUST split them into separate JSON objects. Combine the parent description and the child description. Set the "BOQ Item No." to reflect the split (e.g., "31.i", "31.ii"). Do NOT lump multiple quantities into a single row.

Map the extracted data to these specific keys exactly:
- "BOQ Item No.": The serial number (e.g., "1.", "31.i").
- "Description": The full text description of the work (parent + child if split).
- "Original Quantity": The quantity number (e.g., 32.0). If empty, use null.
- "Original Unit": The unit of measurement (e.g., "Cu.m", "Sq.m", "Quintal").
- "Schedule Item Code": The DSR 1989 Code No. (e.g., "2.8"). If empty, use null.
- "Material Category": Infer the primary material (e.g., "Concrete", "Earthwork", "Steel", "Brick", "Wood").
- "Discipline": Infer one of: "Civil & Sitework", "Structural", "Architectural".
- "Floor / Section": Identify the heading/sub-head from the BOQ.
- "Material / Product": A short specific name (e.g., "Ready mix concrete", "TMT Bar").
- "All Materials Detected": Array of strings of all materials mentioned.
- "Material Confidence": "High", "Medium", or "Low".
- "Grade": Any material grade mentioned.
- "Mix Ratio": Any mix ratio mentioned.
- "Standard / Code Reference": Any Indian Standard codes explicitly mentioned.
- "Classification (Matched)": Create a hierarchical classification string.
- "Length (mm)": Extract from description if present (numeric only).
- "Width (mm)": Extract from description if present (numeric only).
- "Height (mm)": Extract from description if present (numeric only).
- "Thickness (mm)": Extract from description if present (e.g., "12 mm thick" -> 12).
- "Depth (mm)": Extract from description if present (numeric only).
- "Diameter (mm)": Extract from description if present (numeric only).
- "Unit Rate": The rate amount if visible, else null.
- "Total Cost": The total amount if visible, else null.
- "Audit Tag": Use "[OK]" for clean extractions, "[FLAG]" for messy/unclear scans, "[EXCLUDED]" for earth/labour, "[MULTI-SUB]" if you successfully split a parent item.

Respond ONLY with the JSON array. Do NOT extract items outside the range {start} to {end}.
"""

# ---------------------------------------------------------
# 2. HELPER FUNCTIONS
# ---------------------------------------------------------
def generate_with_fallback(client, contents):
    for model_name in FALLBACK_MODELS:
        try:
            response = client.models.generate_content(
                model=model_name,
                contents=contents,
                config=types.GenerateContentConfig(response_mime_type="application/json")
            )
            return response, model_name
        except Exception as e:
            err_str = str(e)
            if "503" in err_str or "404" in err_str:
                print(f"    [!] Model {model_name} failed (503/404). Falling back to next model...")
                continue
            else:
                raise e
    raise RuntimeError("All fallback models are currently unavailable or overloaded.")

def safe_float(val):
    if val is None or val == "":
        return None
    try:
        clean = re.sub(r'[^\d.]', '', str(val))
        return float(clean) if clean else None
    except:
        return None

# FIX 1: Unit Normalization function based on Instruction Sheet
def normalize_unit(raw_unit):
    if not raw_unit: return ""
    u = str(raw_unit).lower().strip()
    if any(x in u for x in ["cum", "cu.m", "m3", "cubic metre", "cubic meter"]): return "cum"
    if any(x in u for x in ["sqm", "sq.m", "m2", "square"]): return "sqm"
    if any(x in u for x in ["rm", "mtr", "metre", "meter"]) or u == "m": return "m"
    if any(x in u for x in ["kg", "kilogram"]): return "kg"
    if any(x in u for x in ["each", "nos", "number"]): return "nos"
    if "quintal" in u: return "quintal"
    if "tonne" in u or "mt" in u: return "tonne"
    return u

# ---------------------------------------------------------
# 3. STATE TRACKING & CHUNKING SETUP
# ---------------------------------------------------------
metadata_done = False
meta_data = {}
chunks = [(1, 7), (8, 14), (15, 21), (22, 28), (29, 35), (36, 42), (43, 49), (50, 56), (57, 61), (62, 64)]
completed_chunks = set()
all_extracted_items = []

print(f"Starting Master Pipeline with {len(api_keys)} Keys and {len(FALLBACK_MODELS)} Models.\n")

for idx, key in enumerate(api_keys):
    if metadata_done and len(completed_chunks) == len(chunks):
        break  
        
    try:
        print(f"\n--- Attempting execution with API Key {idx + 1} of {len(api_keys)} ---")
        client = genai.Client(api_key=key)
        boq_file = client.files.upload(file=PDF_PATH, config={'display_name': f'BoQ_Key_{idx}'})
        
        if not metadata_done:
            print("Extracting Building Metadata...")
            meta_response, used_model = generate_with_fallback(client, [boq_file, meta_prompt])
            meta_data = json.loads(meta_response.text)
            metadata_done = True
            print(f"✔ Metadata extraction successful (using {used_model}).")
            time.sleep(10)

        for start, end in chunks:
            if (start, end) in completed_chunks: continue 
                
            print(f"Extracting Line Items {start} to {end}...")
            chunk_prompt = extract_prompt_template.format(start=start, end=end)
            boq_response, used_model = generate_with_fallback(client, [boq_file, chunk_prompt])
            chunk_items = json.loads(boq_response.text)
            all_extracted_items.extend(chunk_items)
            completed_chunks.add((start, end))
            print(f"✔ Successfully extracted items {start} to {end} (using {used_model}).")
            
            if len(completed_chunks) < len(chunks): time.sleep(10)
            
        print("\n✔ All chunks processed successfully!")
        
    except Exception as e:
        print(f"Error encountered: {e}")
        if "429" in str(e) or "RESOURCE_EXHAUSTED" in str(e):
            print(f"⚠ Rate limit reached on API Key {idx + 1}. Rotating...")
        else:
            break

# ---------------------------------------------------------
# 4. DATA CLEANING, UNIT NORMALIZATION & REGEX 
# ---------------------------------------------------------
current_floor = ""
# Added 2-part Mortar ratios to the mapping
grade_map = {
    "1:5:10": "M5", 
    "1:4:8": "M7.5", 
    "1:3:6": "M10", 
    "1:2:4": "M15", 
    "1:1.5:3": "M20", 
    "1:1:2": "M25",
    "1:6": "Mortar 1:6",
    "1:5": "Mortar 1:5",
    "1:4": "Mortar 1:4",
    "1:3": "Mortar 1:3",
    "1:2": "Mortar 1:2"
}

for item in all_extracted_items:
    raw_u = str(item.get("Original Unit") or "").lower().strip()
    qty = safe_float(item.get("Original Quantity"))
    description = str(item.get("Description") or "")
    
    # --- FIX 1: FILTERED FORWARD FILL 'FLOOR / SECTION' ---
    floor = str(item.get("Floor / Section") or "").strip()
    # Ignore garbage header text
    bad_headers = ["bill of quantities", "p/gen", "modified", "null", "none", ""]
    if floor and not any(bad in floor.lower() for bad in bad_headers):
        current_floor = floor 
    
    item["Floor / Section"] = current_floor

    # --- SPECIAL EDGE CASE: ITEM 24 ---
    if "decimetre" in raw_u:
        qty = qty / 1000.0 if qty else 0
        unit = "cum"
    else:
        unit = normalize_unit(raw_u)
        
    item["Original Quantity"] = qty
    item["Original Unit"] = unit

    # --- FIX 3: AGGRESSIVE IS CODE EXTRACTION (findall) ---
    is_matches = re.findall(r'(IS\s*[:\-]?\s*\d+)', description, re.IGNORECASE)
    if is_matches:
        # Clean and join all found codes (e.g., "IS: 9103, IS: 456")
        cleaned_codes = [re.sub(r'IS\s*[:\-]?\s*', 'IS: ', m.upper()) for m in is_matches]
        item["Standard / Code Reference"] = ", ".join(set(cleaned_codes))

    # --- FIX 2: GRADE MAPPING (Handles both Concrete and Mortar) ---
    mix_ratio = str(item.get("Mix Ratio") or "")
    grade_match = re.search(r'\b(M\s*[-]?\s*\d{1,2}(?:\.5)?)\b', description, re.IGNORECASE)
    
    if grade_match:
        item["Grade"] = grade_match.group(1).upper().replace(' ', '').replace('-', '')
    elif mix_ratio and mix_ratio.lower() not in ["null", "none", ""]:
        ratios = [r.strip().replace(" ", "") for r in mix_ratio.split(',')]
        mapped_grades = []
        for r in ratios:
            if r in grade_map:
                mapped_grades.append(grade_map[r])
        
        if mapped_grades:
            item["Grade"] = ", ".join(mapped_grades)
            
# ---------------------------------------------------------
# 5. SAVE RAW JSON OUTPUTS
# ---------------------------------------------------------
print("\nSaving Output Files...")
with open("output/building_meta.json", "w") as f:
    json.dump(meta_data, f, indent=4)
with open("output/passport.json", "w") as f:
    json.dump(all_extracted_items, f, indent=4)

# ---------------------------------------------------------
# 6. MAP TO EXCEL: GMAP ID, DERIVED MATH, AND AUDIT COMMENTS
# ---------------------------------------------------------
print("Mapping data to Excel Template with Dimensions and Derived Math...")
wb = openpyxl.load_workbook(TEMPLATE_PATH)
ws = wb["Material Passport"]

if ws.max_row >= 4: ws.delete_rows(4, ws.max_row - 3)
headers = [cell.value for cell in ws[3]]

for item in all_extracted_items:
    qty = item.get("Original Quantity")
    unit = item.get("Original Unit")
    mat_cat = str(item.get("Material Category", "")).lower()
    item_no = str(item.get("BOQ Item No.", "")).strip()
    description = str(item.get("Description", "")) 
    
    gmap_id = f"AMP-CBRI-PR-{item_no.replace('.', '-')}" if item_no else ""
    if gmap_id.endswith('-'): gmap_id = gmap_id[:-1]

    all_mats_raw = item.get("All Materials Detected", [])
    all_mats = ", ".join(str(m) for m in all_mats_raw) if isinstance(all_mats_raw, list) else str(all_mats_raw)
    
    density, gwp, db_citation = "", "", ""
    if "concrete" in mat_cat: density, gwp, db_citation = 2400, 0.15, "ICE Database V3.0"
    elif "steel" in mat_cat or "reinforcement" in mat_cat: density, gwp, db_citation = 7850, 2.50, "ICE Database V3.0"
    elif "brick" in mat_cat: density, gwp, db_citation = 1900, 0.24, "ICE Database V3.0"
    elif "wood" in mat_cat or "timber" in mat_cat: density, gwp, db_citation = 650, 0.45, "ICE Database V3.0"
    elif "earth" in mat_cat or "sand" in mat_cat: density, gwp, db_citation = 1600, 0.01, "ICE Database V3.0"

    audit_tag = item.get("Audit Tag", "[OK]")
    if "EXCLUDED" in audit_tag:
        final_comment = f"{audit_tag} negligible embodied material carbon"
    else:
        final_comment = f"{audit_tag} {db_citation}".strip()

    thick_mm = safe_float(item.get("Thickness (mm)"))

    row_dict = {
        "GMAP Id": gmap_id,
        "BOQ Item No.": item_no,
        "Description": description,
        "Original Quantity": qty,
        "Original Unit": unit,
        "Schedule Item Code": item.get("Schedule Item Code"),
        "Material Category": item.get("Material Category"),
        "Discipline": item.get("Discipline"),
        "Floor / Section": item.get("Floor / Section"),
        "Material / Product": item.get("Material / Product"),
        "All Materials Detected": all_mats,
        "Material Confidence": item.get("Material Confidence"),
        "Grade": item.get("Grade"),  # <--- Now safely pulls the Regexed value!
        "Mix Ratio": item.get("Mix Ratio"),
        "Standard / Code Reference": item.get("Standard / Code Reference"), # <--- Now safely pulls the Regexed value!
        "Classification (Matched)": item.get("Classification (Matched)"),
        "Length (mm)": safe_float(item.get("Length (mm)")),
        "Width (mm)": safe_float(item.get("Width (mm)")),
        "Height (mm)": safe_float(item.get("Height (mm)")),
        "Thickness (mm)": thick_mm,
        "Depth (mm)": safe_float(item.get("Depth (mm)")),
        "Diameter (mm)": safe_float(item.get("Diameter (mm)")),
        "Unit Rate": item.get("Unit Rate"),
        "Total Cost": item.get("Total Cost"),
        "Currency": "Rs.",
        "Schedule (DSR/SOR)": "DSR 1989", 
        "Density (kg/m³)": density,
        "GWP / kg (kg CO₂e/kg)": gwp,
        "Comment": final_comment
    }
    

    vol_m3, area_m2, weight_kg = "", "", ""
    derived_q, derived_u, derived_b = "", "", ""

    if qty and float(qty) > 0:
        q = float(qty)
        if unit == "cum":
            vol_m3 = q
            row_dict["Volume (m³)"] = q
        elif unit == "sqm":
            area_m2 = q
            row_dict["Area (m²)"] = q
            if thick_mm:
                derived_q = round(q * (thick_mm / 1000.0), 3)
                derived_u = "Volume (m³)"
                derived_b = f"Area ({q}) * Thickness ({thick_mm}mm)"
                vol_m3 = derived_q 
        elif unit == "m":
            row_dict["Length (m)"] = q
        elif unit == "kg":
            weight_kg = q
            row_dict["Weight (kg)"] = q
        elif unit == "quintal":
            weight_kg = q * 100
            row_dict["Weight (kg)"] = weight_kg
        elif unit == "tonne":
            weight_kg = q * 1000
            row_dict["Weight (kg)"] = weight_kg
        elif unit == "nos":
            row_dict["Count (Nos)"] = q

    row_dict["Derived Quantity"] = derived_q
    row_dict["Derived Quantity Unit"] = derived_u
    row_dict["Derived Quantity Basis"] = derived_b

    if vol_m3 and density and gwp:
        row_dict["Embodied Carbon A1-A3 (kg CO₂e)"] = round(vol_m3 * density * gwp, 2)
    elif weight_kg and gwp:
        row_dict["Embodied Carbon A1-A3 (kg CO₂e)"] = round(weight_kg * gwp, 2)

    row_values = [row_dict.get(col, "") for col in headers]
    ws.append(row_values)

    # --- NEW: APPLY WRAP TEXT & TOP ALIGNMENT ---
    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    # --------------------------------------------

wb.save("output/passport_filled.xlsx")
print("✔ Excel saved: Unit Normalization, Wrap Text, and Special Edge Cases handled.")

# ---------------------------------------------------------
# 7. GENERATE VISUALIZATION CHART
# ---------------------------------------------------------
df_output = pd.DataFrame(all_extracted_items)
if not df_output.empty and "Material Category" in df_output.columns:
    df_plot = df_output["Material Category"].value_counts().reset_index()
    df_plot.columns = ["Material Category", "Count"]

    plt.figure(figsize=(10, 6))
    plt.bar(df_plot["Material Category"], df_plot["Count"], color="#2ecc71")
    plt.title("Distribution of Materials across BoQ Items")
    plt.xlabel("Material Category")
    plt.ylabel("Number of Line Items")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig("output/visualization.png", dpi=300)

print("\nPipeline execution complete! All deliverables generated.")